import os; clear = lambda: os.system('cls'); clear();
import openpyxl as xl; #print("Python Library: ", xl.__name__, xl.__version__)
import pyodbc; import pandas as pd; import warnings; import tabulate as tb;
warnings.filterwarnings("ignore", category=UserWarning)

################################################################################

path_in = r'C:\Users\jordi\Desktop\ARCHIVO\1 - DOCS\PyBcn - OpenPyXL\SCRIPTS\_countries.sql'
path_out = r'C:\Users\jordi\Desktop\ARCHIVO\1 - DOCS\PyBcn - OpenPyXL\XLS\ex1_countries.xlsx'

################################################################################

print("\n***************************************************************************")
print("DB: Northwind; TB: OrderID")
print("***************************************************************************\n")

conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=PC;DATABASE=Northwind;Trusted_Connection=yes;')
query = open(path_in, 'r').read()
df = pd.read_sql(query, conn)

print(df.shape)

print(tb.tabulate(df.head(), headers='keys', tablefmt='psql'))

wb = xl.Workbook()
wb.remove(wb.active)

for country, group in df.groupby('ShipCountry'):

    ws = wb.create_sheet(title=country)

    for col_num, column_title in enumerate(group.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    for row_num, row_data in enumerate(group.itertuples(index=False, name=None), 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

wb.save(path_out) #os.remove("path_out")
conn.close()

print("\nOK 😊👍\n")

print("***************************************************************************\n")

################################################################################

